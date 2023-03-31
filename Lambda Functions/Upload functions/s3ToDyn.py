import boto3
from openpyxl import load_workbook
import datetime

# Initialize DynamoDB client and table name
dynamodb = boto3.resource('dynamodb')
table_name = 'rbkhousingtable'

# Initialize S3 resource
s3 = boto3.resource('s3')


def lambda_handler(event, context):
    # Get the S3 bucket and key from the event
    print("Hello")
    print('Lambda function triggered with event:', event)
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']
    print(key)
    print(bucket)
    # Load the Excel file from S3
    s3.meta.client.download_file(Bucket=bucket, Key=key, Filename="/tmp/file.xlsx")

    print("Object succesfully downloaded")

    obj = load_workbook(filename="/tmp/file.xlsx")

    sheet = obj.active
    data = []
    for row in sheet.iter_rows(min_row=2):
        item = {}
        for i, cell in enumerate(row):
            if type(cell.value) == datetime.datetime:
                cell_value = str(cell.value)
            else:
                cell_value = cell.value
            item[sheet.cell(1, i + 1).value] = cell_value
        data.append(item)

    # Insert the data into DynamoDB
    table = dynamodb.Table(table_name)
    with table.batch_writer() as batch:
        for item in data:
            # skip inserting None values
            print(item.keys())
            cleaned_item = {key: value for key, value in item.items() if value is not None}
            id = cleaned_item['ApplicationId']
            cleaned_item['ApplicationId'] = str(id)
            batch.put_item(Item=cleaned_item)
            print('Writing item to DynamoDB:', cleaned_item)

    print('Data written to DynamoDB table:', table)
    return "Data added to DynamoDB"